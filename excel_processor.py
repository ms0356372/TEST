"""不依賴 GUI 的字串及 Excel 轉換核心。"""
from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.exceptions import CellCoordinatesException


def normalize_value(value: object) -> str:
    """安全正規化 Excel 值；整數型浮點數不保留 .0。"""
    if value is None: return ""
    if isinstance(value, float) and value.is_integer(): return str(int(value))
    return str(value).strip()


def transform_cell(value: object, rules: dict[str, str], input_delimiter: str,
                   output_delimiter: str, sort_order: list[str], keep_unknown: bool = True,
                   remove_duplicates: bool = True) -> str:
    """以原始代碼排序後轉換；未知值依原出現順序置於已知值後方。"""
    text = normalize_value(value)
    if not text: return ""
    if not input_delimiter: raise ValueError("輸入分隔符不可空白")
    tokens = [normalize_value(v) for v in text.split(input_delimiter)]
    tokens = [v for v in tokens if v]
    if remove_duplicates:
        tokens = list(dict.fromkeys(tokens))
    present = set(tokens)
    known = [code for code in sort_order if code in present and code in rules]
    # 未列入 sort_order 的已知規則仍視為已知並依來源順序接續。
    known.extend(v for v in tokens if v in rules and v not in known)
    unknown = [v for v in tokens if v not in rules] if keep_unknown else []
    return output_delimiter.join([rules[v] for v in known] + unknown)


def parse_start_cell(value: str) -> tuple[int, int]:
    """驗證 A1 格式並回傳 (欄, 列)。"""
    try:
        letters, row = coordinate_from_string(value.strip().upper())
        if row < 1: raise ValueError
        return column_index_from_string(letters), row
    except (ValueError, TypeError, CellCoordinatesException):
        raise ValueError("開始儲存格格式錯誤，請輸入例如 E4 或 AA10") from None


def get_sheet_names(path: Path) -> list[str]:
    """僅讀取 workbook 結構取得工作表名稱。"""
    if path.suffix.lower() not in (".xlsx", ".xlsm"): raise ValueError("只支援 .xlsx 與 .xlsm")
    book = load_workbook(path, read_only=True, keep_vba=path.suffix.lower() == ".xlsm")
    try: return book.sheetnames
    finally: book.close()


def preview_rows(path: Path, sheet_name: str, start_cell: str, rules: list[dict[str, str]],
                 input_delimiter: str, output_delimiter: str, unknown_behavior: str,
                 remove_duplicates: bool, limit: int = 20) -> list[tuple[int, object, str]]:
    """唯讀載入並預覽指定起始列後最多 limit 列。"""
    col, row = parse_start_cell(start_cell)
    book = load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    try:
        if sheet_name not in book.sheetnames: raise ValueError("找不到指定工作表")
        ws = book[sheet_name]; mapping = {x["source"]: x["target"] for x in rules}; order = list(mapping)
        return [(r, ws.cell(r, col).value, transform_cell(ws.cell(r, col).value, mapping,
                 input_delimiter, output_delimiter, order, unknown_behavior == "keep", remove_duplicates))
                for r in range(row, min(ws.max_row, row + limit - 1) + 1)]
    finally: book.close()


def process_excel(source: Path, destination: Path, sheet_name: str, start_cell: str,
                  rules: list[dict[str, str]], input_delimiter: str, output_delimiter: str,
                  unknown_behavior: str = "keep", remove_duplicates: bool = True,
                  result_header: str = "轉換結果", create_header: bool = True,
                  progress: Callable[[int, int], None] | None = None) -> dict[str, int]:
    """插入來源右欄、複製來源樣式並另存；絕不允許覆蓋來源檔。"""
    source, destination = source.resolve(), destination.resolve()
    if source == destination: raise ValueError("輸出檔不可覆蓋原始 Excel")
    if source.suffix.lower() != destination.suffix.lower(): raise ValueError("輸出副檔名必須與來源相同")
    col, start_row = parse_start_cell(start_cell)
    book = load_workbook(source, keep_vba=source.suffix.lower() == ".xlsm", data_only=False)
    try:
        if sheet_name not in book.sheetnames: raise ValueError("找不到指定工作表")
        ws = book[sheet_name]; last_row = ws.max_row; result_col = col + 1
        ws.insert_cols(result_col, 1)
        src_letter = ws.cell(1, col).column_letter; dst_letter = ws.cell(1, result_col).column_letter
        ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
        mapping = {x["source"]: x["target"] for x in rules}; order = list(mapping); unknown: set[str] = set()
        for r in range(1, last_row + 1):
            src, dst = ws.cell(r, col), ws.cell(r, result_col)
            if src.has_style:
                dst._style = copy(src._style)
            dst.number_format = src.number_format
        if create_header and start_row > 1: ws.cell(start_row - 1, result_col).value = result_header
        total = max(0, last_row - start_row + 1)
        for done, r in enumerate(range(start_row, last_row + 1), 1):
            value = ws.cell(r, col).value
            tokens = [normalize_value(x) for x in normalize_value(value).split(input_delimiter)] if value is not None else []
            unknown.update(x for x in tokens if x and x not in mapping)
            ws.cell(r, result_col).value = transform_cell(value, mapping, input_delimiter, output_delimiter,
                                                           order, unknown_behavior == "keep", remove_duplicates)
            if progress: progress(done, total)
        destination.parent.mkdir(parents=True, exist_ok=True); book.save(destination)
        return {"processed": total, "unknown": len(unknown)}
    finally: book.close()

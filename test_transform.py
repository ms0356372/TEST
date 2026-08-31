"""核心字串與 Excel 插欄整合測試。"""
import tempfile
import unittest
from pathlib import Path
from openpyxl import Workbook, load_workbook
from excel_processor import parse_start_cell, process_excel, transform_cell

RULES={"99":"一般","02":"噪音","03":"游離輻射"}; ORDER=list(RULES)
class TransformTests(unittest.TestCase):
 def t(self,value,**kw): return transform_cell(value,RULES,kw.pop("input_delimiter",";"),kw.pop("output_delimiter",";"),ORDER,**kw)
 def test_order_one(self): self.assertEqual(self.t("99;03;02"),"一般;噪音;游離輻射")
 def test_order_two(self): self.assertEqual(self.t("03;99;02"),"一般;噪音;游離輻射")
 def test_subset(self): self.assertEqual(self.t("02;99"),"一般;噪音")
 def test_trim(self): self.assertEqual(self.t("99 ; 03 ; 02"),"一般;噪音;游離輻射")
 def test_empty_token(self): self.assertEqual(self.t("99;;03"),"一般;游離輻射")
 def test_duplicate(self): self.assertEqual(self.t("99;99;02"),"一般;噪音")
 def test_unknown_keep(self): self.assertEqual(self.t("99;88;02"),"一般;噪音;88")
 def test_unknown_ignore(self): self.assertEqual(self.t("99;88;02",keep_unknown=False),"一般;噪音")
 def test_blank_none(self): self.assertEqual(self.t(""),"");self.assertEqual(self.t(None),"")
 def test_numeric(self): self.assertEqual(self.t(99),"一般");self.assertEqual(self.t(99.0),"一般")
 def test_other_delimiters(self): self.assertEqual(self.t("99,02",input_delimiter=",",output_delimiter=" / "),"一般 / 噪音")
 def test_no_rules(self): self.assertEqual(transform_cell("88",{},";",";",[],True),"88")
 def test_bad_cell(self):
  self.assertEqual(parse_start_cell("AA10"),(27,10))
  with self.assertRaises(ValueError): parse_start_cell("ABC")

class ExcelTests(unittest.TestCase):
 def test_insert_preserves_source_and_shifts(self):
  with tempfile.TemporaryDirectory() as folder:
   src=Path(folder)/"test_input.xlsx";dst=Path(folder)/"test_output.xlsx";wb=Workbook();ws=wb.active;ws.title="資料"
   ws["E3"]="特殊作業"; values=["99;03;02","02;99","03","99 ; 03 ; 02","99;99;02","99;88;02"]
   for row,value in enumerate(values,4):ws.cell(row,5,value)
   ws["F3"]="部門";ws["F4"]="A部門";wb.save(src)
   result=process_excel(src,dst,"資料","E4",[{"source":k,"target":v} for k,v in RULES.items()],";",";",result_header="特殊作業轉換")
   out=load_workbook(dst);sheet=out["資料"]
   self.assertEqual([sheet.cell(r,5).value for r in range(4,10)],values)
   self.assertEqual([sheet.cell(r,6).value for r in range(3,10)],["特殊作業轉換","一般;噪音;游離輻射","一般;噪音","游離輻射","一般;噪音;游離輻射","一般;噪音","一般;噪音;88"])
   self.assertEqual(sheet["G3"].value,"部門");self.assertEqual(sheet["G4"].value,"A部門");self.assertEqual(result,{"processed":6,"unknown":1});out.close()

if __name__=="__main__": unittest.main(verbosity=2)

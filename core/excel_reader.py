from pathlib import Path
from typing import Any
from core.exceptions import AnalysisError

def read_active_sheet(path:str|Path,sheet_name:str|None=None)->tuple[list[str],list[dict[str,Any]]]:
    from openpyxl import load_workbook
    try: wb=load_workbook(path,read_only=True,data_only=True)
    except Exception as exc: raise AnalysisError(f"Excel 無法讀取：{exc}") from exc
    try:
        ws=wb[sheet_name] if sheet_name else wb.active
        values=ws.iter_rows(values_only=True); first=next(values,None)
        if first is None:raise AnalysisError("Excel 沒有表頭。")
        headers=[str(v).strip() if v is not None else "" for v in first]
        duplicates={h for h in headers if h and headers.count(h)>1}
        if duplicates:raise AnalysisError("表頭重複："+"、".join(sorted(duplicates)))
        rows=[]
        for values_row in values:
            if all(v is None or str(v).strip()=="" for v in values_row):continue
            rows.append({h:v for h,v in zip(headers,values_row) if h})
        return headers,rows
    finally:wb.close()

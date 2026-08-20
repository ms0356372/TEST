from pathlib import Path
from typing import Any, Iterable, Sequence


def write_analysis_workbook(
    path: str | Path,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    header_fill_ranges: Sequence[tuple[int, int, str]] | None = None,
) -> Path:
    """Write a consistently styled analysis workbook.

    ``header_fill_ranges`` contains inclusive, one-based column ranges.  The
    default preserves the original Framingham output styling.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    base_font = Font(name="標楷體", size=10)
    center = Alignment(horizontal="center", vertical="center")
    ranges = header_fill_ranges or ((1, len(headers), "F2F2F2"),)
    risk_fills = {
        "中度": PatternFill("solid", fgColor="D6DCE4"),
        "高度": PatternFill("solid", fgColor="FFFF00"),
        "極高": PatternFill("solid", fgColor="FF0000"),
    }
    higher_fill = PatternFill("solid", fgColor="FFFF00")

    target = Path(path)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "分析結果"
    worksheet.append(list(headers))
    for row in rows:
        worksheet.append(list(row))
    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = base_font
            cell.alignment = center
    for start, end, color in ranges:
        fill = PatternFill("solid", fgColor=color)
        for column in range(start, end + 1):
            worksheet.cell(1, column).fill = fill

    # Framingham-specific result colors only apply when those columns exist.
    if len(headers) == 23 and headers[-2:] == ("評估十年內風險程度", "相較同年齡發生率"):
        for row_number in range(2, worksheet.max_row + 1):
            risk_level = worksheet.cell(row_number, 22).value
            if risk_level in risk_fills:
                worksheet.cell(row_number, 22).fill = risk_fills[risk_level]
            if worksheet.cell(row_number, 23).value == "較高":
                worksheet.cell(row_number, 23).fill = higher_fill
    for column in range(1, worksheet.max_column + 1):
        width = max((len(str(worksheet.cell(row, column).value or "")) for row in range(1, worksheet.max_row + 1)), default=8)
        worksheet.column_dimensions[get_column_letter(column)].width = min(max(width + 2, 10), 40)
    worksheet.freeze_panes = "A2"
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(target)
    return target

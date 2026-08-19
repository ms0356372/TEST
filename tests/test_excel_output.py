import tempfile,unittest
from pathlib import Path
try:
 from openpyxl import Workbook,load_workbook
except ImportError:
 Workbook=load_workbook=None
from analyses.framingham_v1 import FraminghamV1Analysis,OUTPUT_HEADERS
@unittest.skipIf(Workbook is None,"openpyxl 未安裝")
class Integration(unittest.TestCase):
 def test_complete_workflow_and_styles(self):
  with tempfile.TemporaryDirectory() as d:
   source=Path(d)/"source.xlsx";wb=Workbook();ws=wb.active;headers=["工號","姓名","廠別","部門","性別","年齡","*收縮壓","*舒張壓","*膽固醇","HDL-C","*抽菸","既往病史"];ws.append(headers)
   cases=[("1","正常男性","A","部","男",40,120,80,180,50,"從未吸菸",""),("2","正常女性","A","部","女",40,120,80,180,50,"從未吸菸",""),("3","高風險男性","A","部","男",70,170,105,300,30,"有","糖尿病"),("4","高風險女性","A","部","女",70,170,105,300,30,"有","糖尿病"),("5","字串數值","A","部","男","50","140 ","90","240","34","有",""),("6","部分未檢","A","部","女",50,"未檢",80,200,50,"從未吸菸","")]
   for row in cases:ws.append(row)
   wb.save(source);before=source.read_bytes();logs=[];out=FraminghamV1Analysis().run({"總表":str(source)},d,logs.append);self.assertEqual(source.read_bytes(),before)
   result=load_workbook(out);sheet=result.active;self.assertEqual(sheet.max_column,23);self.assertEqual(tuple(c.value for c in sheet[1]),OUTPUT_HEADERS)
   for row in sheet.iter_rows():
    for c in row:self.assertEqual(c.font.name,"標楷體");self.assertEqual(c.font.sz,10);self.assertEqual(c.alignment.horizontal,"center");self.assertEqual(c.alignment.vertical,"center")
   self.assertEqual(sheet["A1"].fill.fgColor.rgb[-6:],"F2F2F2");self.assertEqual(sheet.cell(4,22).fill.fgColor.rgb[-6:],"FF0000");self.assertEqual(sheet.cell(4,23).fill.fgColor.rgb[-6:],"FFFF00");self.assertTrue(any("未檢" not in x and "無法轉換" in x for x in logs))

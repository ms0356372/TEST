import tempfile
import unittest
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = load_workbook = None

from analyses.middle_aged import (
    DISEASE_HEADERS, OUTPUT_HEADERS, Q1, Q21_SOURCE, Q22_SOURCE, Q4, Q5, Q6,
    Q71, Q72, Q73, Q21_OUTPUT, Q22_OUTPUT, REQUIRED_HEADERS, MiddleAgedAnalysis,
)
from core.exceptions import AnalysisError


@unittest.skipIf(Workbook is None, "openpyxl 未安裝")
class MiddleAgedExcelIntegrationTests(unittest.TestCase):
    def _valid_answers(self, name="正常案例"):
        values = {header: "" for header in REQUIRED_HEADERS}
        values.update({
            "工號":"A001", "姓名":name, "性別":"女", "廠別":"一廠", "部門":"生產", "課別":"甲課",
            Q1:"0分表示目前完全無法工作，10分表示目前工作能力最佳)請選擇最適合的分數:10",
            Q21_OUTPUT:"很好", Q22_OUTPUT:"很好", Q4:"沒有任何影響/我沒有任何疾病",
            Q5:"0天", Q6:"應該可以", Q71:"總是", Q72:"總是", Q73:"總是",
        })
        return values

    def _write_source(self, path, headers, records):
        workbook = Workbook(); sheet = workbook.active; sheet.append(headers)
        for record in records: sheet.append([record.get(header) for header in headers])
        workbook.save(path)

    def test_age_column_full_workflow_and_styles(self):
        with tempfile.TemporaryDirectory() as folder:
            root = Path(folder); source = root / "test_middle_aged_input.xlsx"
            records = []
            highest = self._valid_answers("最高分案例"); highest["年齡"] = 50; records.append(highest)
            lowest = self._valid_answers("低分案例"); lowest.update({"年齡":55,Q1:"分數:0",Q21_OUTPUT:"很不好",Q22_OUTPUT:"很不好",Q4:"自己覺得完全不能工作",Q5:"100-365天",Q6:"不太可能",Q71:"從不",Q72:"從不",Q73:"從不",DISEASE_HEADERS[0]:"背部",DISEASE_HEADERS[1]:"類風濕性關節炎",DISEASE_HEADERS[2]:"高血壓",DISEASE_HEADERS[3]:"肺結核",DISEASE_HEADERS[4]:"精神疾病或嚴重心理健康問題"}); records.append(lowest)
            missing = self._valid_answers("部分答案缺失"); missing.update({"年齡":60,Q21_OUTPUT:""}); records.append(missing)
            no_disease = self._valid_answers("完全沒有疾病案例"); no_disease["年齡"] = 45; records.append(no_disease)
            colon_case = self._valid_answers("冒號排除案例"); colon_case.update({"年齡":50,DISEASE_HEADERS[8]:"其他皮膚疾病：12、過敏性皮疹或紅斑、其他疹子：12"}); records.append(colon_case)
            headers = [*REQUIRED_HEADERS, Q21_OUTPUT, Q22_OUTPUT, "年齡"]
            self._write_source(source, headers, records); original = source.read_bytes(); logs=[]
            output = MiddleAgedAnalysis().run({"中高齡原稿":str(source)}, str(root), logs.append)
            self.assertEqual(source.read_bytes(), original)
            workbook = load_workbook(output); sheet = workbook.active
            self.assertEqual(sheet.max_column, 41); self.assertEqual(tuple(cell.value for cell in sheet[1]), OUTPUT_HEADERS)
            self.assertEqual(sheet.cell(2,9).value, "很好"); self.assertEqual(sheet.cell(2,10).value, "很好")
            self.assertEqual(sheet.cell(2,32).value, 10)
            self.assertEqual(sheet.cell(2,38).value, 49); self.assertEqual(sheet.cell(2,39).value, "優")
            self.assertEqual(sheet.cell(3,33).value, 1); self.assertEqual(sheet.cell(3,38).value, 7)
            self.assertIsNone(sheet.cell(4,38).value); self.assertEqual(sheet.cell(5,33).value, 7)
            # 三個皮膚疾病項目 -> AG=3；其餘皆最高分 -> AL=45、良。
            self.assertEqual(sheet.cell(6,33).value, 3); self.assertEqual(sheet.cell(6,38).value, 45)
            self.assertEqual(sheet.cell(6,39).value, "良")
            self.assertEqual(sheet.cell(6,40).value, "能勝任所從事的工作")
            self.assertEqual(sheet.cell(6,41).value, "支持其工作適能")
            for column in range(1, 42):
                expected = "F2F2F2" if column <= 30 else "E2EFDA" if column <= 37 else "D9E1F2"
                self.assertEqual(sheet.cell(1,column).fill.fgColor.rgb[-6:], expected)
            for row in sheet.iter_rows():
                for cell in row:
                    self.assertEqual(cell.font.name,"標楷體"); self.assertEqual(cell.font.sz,10)
                    self.assertEqual(cell.alignment.horizontal,"center"); self.assertEqual(cell.alignment.vertical,"center")
            self.assertTrue(any("第 2 題答案無法辨識" in message for message in logs))

    def test_roc_dates_and_invalid_date_continue(self):
        with tempfile.TemporaryDirectory() as folder:
            root=Path(folder); source=root/"test_middle_aged_input.xlsx"
            valid=self._valid_answers("日期計算案例"); valid.update({"出生日期":"087/01/01","體檢日期":"115/08/19"})
            invalid=self._valid_answers("日期異常"); invalid.update({"出生日期":"abc","體檢日期":"1150819"})
            headers=[*REQUIRED_HEADERS,Q21_OUTPUT,Q22_OUTPUT,"出生日期","體檢日期"]
            self._write_source(source,headers,[valid,invalid]); logs=[]
            output=MiddleAgedAnalysis().run({"中高齡原稿":str(source)},str(root),logs.append)
            sheet=load_workbook(output).active
            self.assertEqual(sheet.cell(2,4).value,28); self.assertIsNone(sheet.cell(3,4).value)
            self.assertTrue(any("出生日期或體檢日期無法解析" in message for message in logs))

    def test_duplicate_question_headers_block_analysis(self):
        with tempfile.TemporaryDirectory() as folder:
            root=Path(folder); source=root/"test_middle_aged_input.xlsx"
            record=self._valid_answers(); duplicate=Q21_SOURCE+"(第二個候選)"; record[duplicate]="好"
            headers=[*REQUIRED_HEADERS,Q21_OUTPUT,duplicate,Q22_OUTPUT,"年齡"]; record["年齡"]=50
            self._write_source(source,headers,[record]); logs=[]
            with self.assertRaisesRegex(AnalysisError,"偵測到多個包含以下文字的表頭"):
                MiddleAgedAnalysis().run({"中高齡原稿":str(source)},str(root),logs.append)
            self.assertTrue(any("欄：" in message and Q21_SOURCE in message for message in logs))

    def test_missing_template_is_user_facing_error(self):
        with tempfile.TemporaryDirectory() as folder:
            with self.assertRaisesRegex(AnalysisError,"尚未匯入「中高齡原稿」"):
                MiddleAgedAnalysis().run({},folder)

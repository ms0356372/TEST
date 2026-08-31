"""啟動及打包後共用的自我檢查。"""
from __future__ import annotations

import tempfile
import traceback
from pathlib import Path

from openpyxl import Workbook, load_workbook
from config_manager import get_app_dir, load_config
from excel_processor import transform_cell
from logger import get_logger
from version import APP_VERSION


def run_self_check(callback=None) -> list[dict[str, str]]:
    """逐項執行檢查並回報 PASS/WARNING/FAIL；callback 可供 splash 即時更新。"""
    results: list[dict[str, str]] = []; log = get_logger()
    def check(name, action):
        try:
            status, message = action()
        except Exception as exc:
            status, message = "FAIL", str(exc); log.error("%s failed\n%s", name, traceback.format_exc())
        item = {"name": name, "status": status, "message": message}; results.append(item)
        log.info("[%s] %s: %s", status, name, message)
        if callback: callback(item, len(results))
    check("程式版本", lambda: ("PASS", f"v{APP_VERSION}"))
    check("設定資料夾", lambda: _check_directory())
    check("設定檔", lambda: _check_config())
    check("Excel 模組", lambda: ("PASS", "openpyxl 正常"))
    check("核心轉換", lambda: _check_transform())
    check("寫入權限", lambda: _check_write())
    check("暫存 Excel 與中文", lambda: _check_excel())
    check("未知代碼與重複值", lambda: _check_edges())
    return results


def _check_directory():
    get_app_dir().mkdir(parents=True, exist_ok=True); return "PASS", str(get_app_dir())


def _check_config():
    _, repaired = load_config(); return ("WARNING", "設定損壞，已備份並重建") if repaired else ("PASS", "正常")


def _check_transform():
    rules = {"99": "一般", "02": "噪音", "03": "游離輻射"}
    assert transform_cell("03;99;02", rules, ";", ";", list(rules)) == "一般;噪音;游離輻射"
    return "PASS", "排序正確"


def _check_write():
    with tempfile.NamedTemporaryFile(prefix="excel_tool_", delete=True) as stream: stream.write(b"ok")
    return "PASS", "暫存目錄可寫入"


def _check_excel():
    with tempfile.TemporaryDirectory() as folder:
        path = Path(folder) / "中文測試.xlsx"; book = Workbook(); book.active["A1"] = "中文測試"; book.save(path)
        loaded = load_workbook(path, read_only=True)
        try: assert loaded.active["A1"].value == "中文測試"
        finally: loaded.close()
    return "PASS", "建立、儲存、讀取正常"


def _check_edges():
    rules = {"99": "一般", "02": "噪音"}
    assert transform_cell("99;88;99;02", rules, ";", ";", list(rules)) == "一般;噪音;88"
    assert transform_cell("99;88", rules, ";", ";", list(rules), False) == "一般"
    return "PASS", "處理正確"


def write_result_file(path: Path, results: list[dict[str, str]]) -> None:
    """供 windowed EXE 將結果可靠交回 BAT。"""
    passed = not any(x["status"] == "FAIL" for x in results)
    lines = ["PASS" if passed else "FAIL"] + [f'[{x["status"]}] {x["name"]}: {x["message"]}' for x in results]
    path.parent.mkdir(parents=True, exist_ok=True); path.write_text("\n".join(lines), encoding="utf-8")
